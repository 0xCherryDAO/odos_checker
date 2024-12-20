import random
from asyncio import sleep, gather, create_task

from loguru import logger
from rich.console import Console
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config import MOBILE_PROXY, PAUSE_BETWEEN_WALLETS, ROTATE_IP
from src.checker.odos_checker import Checker
from src.utils.data.helper import proxies
from src.utils.proxy_manager import Proxy


async def process_wallet(address: str, proxy_index: int) -> list:
    proxy = await prepare_proxy(proxies[proxy_index])
    odos = Checker(
        address=address,
        proxy=proxy
    )
    tokens = await odos.check_for_tokens()
    row = [
        address,
        tokens,
    ]
    return row


async def process_checker(addresses: list[str]):
    workbook = Workbook()
    console = Console()
    sheet = workbook.active
    sheet.title = "Odos Data"

    headers = [
        "address",
        "total tokens",
    ]
    sheet.append(headers)
    tasks = []
    with console.status("[bold green]Checking wallets..."):
        for index, address in enumerate(addresses):
            console.log(f'Checking {address}...')
            proxy_index = index % len(proxies)
            tasks.append(create_task(process_wallet(address, proxy_index)))
            time_to_sleep = random.uniform(PAUSE_BETWEEN_WALLETS[0], PAUSE_BETWEEN_WALLETS[1])
            if time_to_sleep != 0:
                console.log(f'Sleeping {time_to_sleep:.2f} seconds...')
                await sleep(time_to_sleep)
        results = await gather(*tasks)

    for row in results:
        sheet.append(row)
        console.log(f"Wallet {row[0]} checked")

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = max_length + 2

    workbook.save("odos_data.xlsx")
    logger.success("Excel file 'odos_data.xlsx' has been saved!")


async def prepare_proxy(proxy: str) -> Proxy | None:
    if proxy:
        change_link = None
        if MOBILE_PROXY:
            proxy_url, change_link = proxy.split('|')
        else:
            proxy_url = proxy

        proxy = Proxy(proxy_url=f'http://{proxy_url}', change_link=change_link)

        if ROTATE_IP and MOBILE_PROXY:
            await proxy.change_ip()

        return proxy
